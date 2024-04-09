import openpyxl
import pyodbc
import os
import requests
from bs4 import BeautifulSoup
import datetime
import random
from re import sub


class updatebrecords:
    def __init__(self):

        # WILL OUTPUT ALL ERRORS ENCOUNTERED AFTER ATTEMPTING TO LAUNCH PROGRAM (IF ISSUES ARE FOUND)
        self.erroroutputmsg = "The following issue(s) have been encountered:\n"
        self.errorencountered = False

        try:
            self.cnxn = pyodbc.connect("DRIVER={SQL Server}; Server=localhost\SQLEXPRESS; Database=master; Trusted_Connection=True;")  # CONNECTION STRING
            self.cursor = self.cnxn.cursor()  # cursor WILL BE USED TO RETRIEVE A RESULT FROM AN SQL STATEMENT
        except pyodbc.DatabaseError as connectionerror:
            self.problemsencountered = self.erroroutputmsg + "  - {}".format(connectionerror) + "\n"
            self.errorencountered = True

    def checkspreadsheets(self):

        # STARTS ADDING NEW PARTS TO THE DATABASE TABLE BikeParts

        path = os.getcwd() + "/Part_Spreadsheet.xlsx"  # FINDS DIRECTORY FOR Part_Spreadsheet.xlsx
        file = openpyxl.load_workbook(path).active

        print("Checking Part_Spreadsheet.xlsx for new parts added.")

        sqlstatement = f'SELECT PartID FROM BikeParts ORDER BY PartID ASC'
        self.cursor.execute(sqlstatement)  # RETRIEVES ALL PARTS ALREADY IN THE DATABASE
        partsalreadyindb = self.cursor.fetchall()
        partsalreadyindb = [item[0] for item in partsalreadyindb]

        for i in range(1, int(file.max_row)):  # LOOPS THROUGH ALL ROWS IN Part_Spreadsheet.xlsx

            # if found_index == -1:  # PART ALREADY STORED IN BikeComponents.
            if file.cell(row=i, column=3).value in partsalreadyindb:  # PART ALREADY STORED IN BikeComponents.
                print(f'Already stored: {file.cell(row=i, column=1).value}. Starting web scraping.')
                updatebrecords().checkprices(file.cell(row=i, column=1).value, file.cell(row=i, column=3).value)  # WEB SCRAPING PRICES FOR THE CURRENT PART
            else:  # PART NEEDS TO BE ADDED TO BikeComponents.
                try:
                    sqlstatement = f'INSERT INTO BikeParts VALUES (\'{file.cell(row=i, column=3).value}\', \'{file.cell(row=i, column=1).value}\', \'{file.cell(row=i, column=4).value}\', \'{file.cell(row=i, column=2).value}\')'
                    self.cursor.execute(sqlstatement)
                    self.cursor.commit()
                    print(f'Added: {file.cell(row=i, column=1).value}. Starting web scraping')
                    updatebrecords().checkprices(file.cell(row=i, column=1).value, file.cell(row=i, column=3).value)  # WEB SCRAPING PRICES FOR THE CURRENT PART
                except pyodbc.DatabaseError as connectionerror:
                    self.problemsencountered = self.erroroutputmsg + "  - {}".format(connectionerror) + "\n"
                    self.errorencountered = True

        if self.errorencountered:  # OUTPUTS ERRORS IF THERE WERE ANY
            print(self.problemsencountered)

    def checkprices(self, partname, partid):

        url = f'https://www.bing.com/shop?q={partname}&FORM=SHOPTB'

        user_agent_list = [  # LIST OF RANDOM USER AGENTS TO SIMULATE A REAL USER ACCESSING EDGE
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/13.1.1 Safari/605.1.15',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:77.0) Gecko/20100101 Firefox/77.0',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.97 Safari/537.36',
            'Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:77.0) Gecko/20100101 Firefox/77.0',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/83.0.4103.97 Safari/537.36',
        ]

        header = requests.utils.default_headers()
        header.update({'User-Agent': random.choice(user_agent_list)})  # RANDOM USER AGENT SELECTED

        sourcecode = requests.get(url, headers=header)  # SENDS GET REQUEST TO THE SPECIFIED URL (url)
        sourcecode = BeautifulSoup(sourcecode.text, "lxml")  # CREATES A BeautifulSoup OBJECTS FROM THE HTML CONTENT USING THE SPECIFIED PARSER (lxml)

        listings = sourcecode.find_all('div', {"class": 'slide'})  # FINDS DIV FOR EACH PART LISTING
        bestmatcharray = []  # STORES THE INFORMATION OF THE PART LISTING INFORMATION TO STORE IN PriceChanges DATABASE TABLE
        bestmatch = 0  # SETS lowestprice TO A VERY HIGH INITIAL VALUE THAT WILL BE REPLACED BY A LOWER PRICE IN THE i LOOP
        for i in range(0, len(listings)):
            price = listings[i].find('div', {"class": 'br-price'})  # FINDS PRICE
            name = listings[i].find('span', title=True)  # FINDS LISTED PART NAME
            retailer = listings[i].find('span', {"class": 'br-offSlrTxt'})  # FINDS RETAILER NAME
            link = listings[i].find('a', href=True)  # FINDS WEBSITE LINK
            if price is not None and name is not None and retailer is not None and link is not None:  # CHECKS THAT NO INFORMATION IS EMPTY
                jarowinklerdistance = updatebrecords.jarowinkler(self, partname, name['title'])  # COMPARES PART LISTING NAME TO THE ACTUAL PART NAME AND RETURNS VALUE 0 TO 1
                if jarowinklerdistance > bestmatch:
                    bestmatch = jarowinklerdistance  # UPDATES lowestprice TO THE LOWEST PRICE FOUND
                    for element in retailer:
                        retailer = element.get_text()  # EXTRACTS RETAILER NAME
                    for element in price:
                        price = sub(r'[^\d.]', '', element.get_text())  # EXTRACTS PRICE
                    bestmatcharray = [price, name['title'], retailer, link['href']]  # SETS bestmatcharray TO VALID MATCH

        if bestmatcharray:
            sqlstatement = f'INSERT INTO PriceChanges VALUES (\'{bestmatcharray[0]}\', \'{datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")}\', \'{bestmatcharray[2]}\', \'{bestmatcharray[3]}\', \'{partid}\')'
            print(sqlstatement)
            self.cursor.execute(sqlstatement)  # INSERTS THE INFORMATION STORED IN bestmatcharray TO THE PriceChanges DATABASE TABLE
            self.cursor.commit()

    def jarowinkler(self, s: str, t: str) -> float:
        m, n = len(s), len(t)
        match_distance = max(m, n) // 2 - 1
        s_matches = [False] * m
        t_matches = [False] * n
        matches = 0
        transpositions = 0
        for i in range(m):
            start = max(0, i - match_distance)
            end = min(i + match_distance + 1, n)
            for j in range(start, end):
                if not t_matches[j] and s[i] == t[j]:
                    s_matches[i] = True
                    t_matches[j] = True
                    matches += 1
                    break
        if matches == 0:
            return 0.0
        k = 0
        for i in range(m):
            if s_matches[i]:
                while not t_matches[k]:
                    k += 1
                if s[i] != t[k]:
                    transpositions += 1
                k += 1
        similarity = (matches / m + matches / n + (matches - transpositions // 2) / matches) / 3
        prefixlength = 0
        for i in range(min(4, min(m, n))):
            if s[i] == t[i]:
                prefixlength += 1
            else:
                break
        return similarity + prefixlength * 0.1 * (1 - similarity)


updatebrecords().checkspreadsheets()
